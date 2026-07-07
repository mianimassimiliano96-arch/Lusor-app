import os
import sys
import json
import datetime
import copy
import zipfile
import tempfile
import shutil
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.filechooser import FileChooserListView
from kivy.clock import Clock
from kivy.utils import platform
from kivy.metrics import dp
from models import Edificio, Stanza, Apparecchio

PIANI = ["Terra", "Primo", "Secondo", "Terzo", "Quarto", "Seminterrato", "Sottotetto", "Altro"]
DESTINAZIONI = ["Ufficio", "Aula/Corsi", "Corridoio", "Magazzino", "Bagno", "Sala riunioni", "Reception", "Cucina", "Altro"]
TIPI_CT = ["Cartongesso", "Grigliato metallico", "Lamellare", "Pannelli fonoassorbenti", "Doghe in legno", "PVC", "Nessuno / A vista", "Altro"]
TIPOLOGIE_APP = ["LED Lineare", "LED Panel", "LED Downlight", "LED Strip", "LED Flood", "Fluorescente", "Alogena", "Incandescenza", "Altro"]
INSTALLAZIONI = ["Soffitto", "Parete", "Pensile", "A sospensione", "A pavimento", "Altro"]
ACCENSIONI = ["Interruttore", "Dimmer", "Sensore presenza", "Crepuscolare", "Timer", "Domotica", "Altro"]


class MessaggioPopup(Popup):
    def __init__(self, titolo, testo, **kwargs):
        super().__init__(title=titolo, size_hint=(0.7, 0.35), **kwargs)
        layout = BoxLayout(orientation="vertical", padding=20, spacing=15)
        layout.add_widget(Label(text=testo, halign="center"))
        btn = Button(text="OK", size_hint_y=0.4, on_press=self.dismiss)
        layout.add_widget(btn)
        self.add_widget(layout)


class ConfermaPopup(Popup):
    def __init__(self, titolo, testo, callback, **kwargs):
        super().__init__(title=titolo, size_hint=(0.7, 0.35), **kwargs)
        self.callback = callback
        layout = BoxLayout(orientation="vertical", padding=20, spacing=15)
        layout.add_widget(Label(text=testo, halign="center"))
        btn_row = BoxLayout(spacing=10, size_hint_y=0.4)
        btn_si = Button(text="Si", on_press=self._conferma)
        btn_no = Button(text="No", on_press=self.dismiss)
        btn_row.add_widget(btn_si)
        btn_row.add_widget(btn_no)
        layout.add_widget(btn_row)
        self.add_widget(layout)

    def _conferma(self, *args):
        self.dismiss()
        if self.callback:
            self.callback()


class HomeScreen(Screen):
    def on_enter(self, *args):
        app = App.get_running_app()
        if "nome_input" in self.ids:
            self.ids.nome_input.text = app.edificio.nome
            self.ids.indirizzo_input.text = app.edificio.indirizzo
            self.ids.data_input.text = app.edificio.data_rilievo
        else:
            Clock.schedule_once(lambda *x: self.on_enter(), 0.2)

    def salva_edificio(self):
        app = App.get_running_app()
        app.edificio.nome = self.ids.nome_input.text.strip()
        app.edificio.indirizzo = self.ids.indirizzo_input.text.strip()
        app.edificio.data_rilievo = self.ids.data_input.text.strip() or datetime.datetime.now().strftime("%d/%m/%Y")
        app.salva_json()
        self.manager.current = "stanze"

    def esporta_risultati(self):
        app = App.get_running_app()
        if not app.edificio.stanze:
            app.mostra_messaggio("Attenzione", "Nessuna stanza da esportare.")
            return
        self.manager.current = "export"


class RoomListScreen(Screen):
    def on_enter(self, *args):
        if "lista_stanze" in self.ids:
            self.carica_lista()
        else:
            Clock.schedule_once(lambda *x: self.on_enter(), 0.2)

    def carica_lista(self):
        app = App.get_running_app()
        grid = self.ids.lista_stanze
        grid.clear_widgets()
        if not app.edificio.stanze:
            grid.add_widget(Label(text="Nessuna stanza. Tocca + per aggiungere.", size_hint_y=None, height=dp(40)))
            return
        for i, s in enumerate(app.edificio.stanze):
            n_app = sum(a.quantita for a in s.apparecchi)
            n_foto = len(s.foto_paths)
            card = BoxLayout(orientation="vertical", size_hint_y=None, height=dp(80), padding=dp(8), spacing=dp(2))
            card.bind(minimum_height=lambda *x: setattr(card, "height", max(dp(80), card.minimum_height)))
            info = f"{s.id_locale}  |  {s.piano}  |  {s.destinazione}  |  App: {n_app}  |  Foto: {n_foto}"
            lbl = Label(text=info, halign="left", valign="middle", size_hint_y=None, height=dp(40), font_size=dp(14))
            lbl.bind(texture_size=lambda *x: setattr(lbl, "height", max(dp(40), lbl.texture_size[1])))
            btn_row = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(8))
            btn_mod = Button(text="Modifica", size_hint_x=0.5)
            btn_mod.bind(on_press=lambda *args, idx=i: self.modifica_stanza(idx))
            btn_dup = Button(text="Duplica", size_hint_x=0.25)
            btn_dup.bind(on_press=lambda *args, idx=i: self.duplica_stanza(idx))
            btn_del = Button(text="X", size_hint_x=0.25, background_color=(0.8, 0.2, 0.2, 1))
            btn_del.bind(on_press=lambda *args, idx=i: self.elimina_stanza(idx))
            for w in [btn_mod, btn_dup, btn_del]:
                btn_row.add_widget(w)
            card.add_widget(lbl)
            card.add_widget(btn_row)
            grid.add_widget(card)

    def nuova_stanza(self):
        self.manager.current = "stanza_detail"
        self.manager.get_screen("stanza_detail").reset()
        self.manager.get_screen("stanza_detail").index = -1

    def modifica_stanza(self, idx):
        self.manager.current = "stanza_detail"
        self.manager.get_screen("stanza_detail").carica_stanza(idx)

    def duplica_stanza(self, idx):
        app = App.get_running_app()
        copia = copy.deepcopy(app.edificio.stanze[idx])
        copia.id_locale = copia.id_locale + " (copia)"
        copia.foto_paths = []
        app.edificio.stanze.append(copia)
        self.carica_lista()

    def elimina_stanza(self, idx):
        app = App.get_running_app()
        stanza = app.edificio.stanze[idx]

        def conferma():
            app.edificio.stanze.pop(idx)
            self.carica_lista()

        popup = ConfermaPopup("Elimina", f"Eliminare '{stanza.id_locale}'?", conferma)
        popup.open()


class FixturePopup(Popup):
    def __init__(self, apparecchio=None, callback=None, **kwargs):
        super().__init__(title="Apparecchio Illuminante", size_hint=(0.9, 0.7), **kwargs)
        self.callback = callback
        a = apparecchio or Apparecchio()
        layout = BoxLayout(orientation="vertical", padding=15, spacing=8)
        scroll = ScrollView()
        form = GridLayout(cols=2, spacing=8, size_hint_y=None, padding=[0, 0, 0, dp(40)])
        form.bind(minimum_height=form.setter("height"))

        form.add_widget(Label(text="Tipologia:"))
        self.tipologia = Spinner(text=a.tipologia or "LED Panel", values=TIPOLOGIE_APP, size_hint_x=0.6)
        form.add_widget(self.tipologia)

        form.add_widget(Label(text="Potenza (W):"))
        self.potenza = TextInput(text=a.potenza, size_hint_x=0.6, input_filter="float")
        form.add_widget(self.potenza)

        form.add_widget(Label(text="Installazione:"))
        self.installazione = Spinner(text=a.installazione or "Soffitto", values=INSTALLAZIONI, size_hint_x=0.6)
        form.add_widget(self.installazione)

        form.add_widget(Label(text="Altezza installaz. (m):"))
        self.altezza_inst = TextInput(text=a.altezza_installazione, size_hint_x=0.6, input_filter="float")
        form.add_widget(self.altezza_inst)

        form.add_widget(Label(text="Accensione:"))
        self.accensione = Spinner(text=a.accensione or "Interruttore", values=ACCENSIONI, size_hint_x=0.6)
        form.add_widget(self.accensione)

        form.add_widget(Label(text="Quantita:"))
        self.quantita = TextInput(text=str(a.quantita), size_hint_x=0.6, input_filter="int")
        form.add_widget(self.quantita)

        scroll.add_widget(form)
        layout.add_widget(scroll)

        btn_row = BoxLayout(size_hint_y=0.1, spacing=10)
        btn_save = Button(text="Salva", on_press=self._salva)
        btn_cancel = Button(text="Annulla", on_press=self.dismiss)
        btn_row.add_widget(btn_save)
        btn_row.add_widget(btn_cancel)
        layout.add_widget(btn_row)
        self.add_widget(layout)

    def _salva(self, *args):
        a = Apparecchio(
            tipologia=self.tipologia.text,
            potenza=self.potenza.text,
            installazione=self.installazione.text,
            altezza_installazione=self.altezza_inst.text,
            accensione=self.accensione.text,
            quantita=int(self.quantita.text) if self.quantita.text.isdigit() else 1,
        )
        self.dismiss()
        if self.callback:
            self.callback(a)


class RoomDetailScreen(Screen):
    index = -1

    def reset(self):
        self.ids.id_locale.text = ""
        self.ids.piano.text = PIANI[0]
        self.ids.altezza.text = "3.0"
        self.ids.destinazione.text = DESTINAZIONI[0]
        self.ids.controsoffitto.text = TIPI_CT[0]
        self.ids.apparecchi_list.clear_widgets()
        self.ids.foto_list.clear_widgets()
        self.foto_paths = []
        self.apparecchi = []
        self.index = -1

    def carica_stanza(self, idx):
        app = App.get_running_app()
        s = app.edificio.stanze[idx]
        self.index = idx
        self.ids.id_locale.text = s.id_locale
        self.ids.piano.text = s.piano if s.piano in PIANI else PIANI[0]
        self.ids.altezza.text = s.altezza
        self.ids.destinazione.text = s.destinazione if s.destinazione in DESTINAZIONI else DESTINAZIONI[0]
        self.ids.controsoffitto.text = s.controsoffitto if s.controsoffitto in TIPI_CT else TIPI_CT[0]
        self.apparecchi = copy.deepcopy(s.apparecchi)
        self.foto_paths = list(s.foto_paths)
        self.aggiorna_apparecchi()
        self.aggiorna_foto()

    def aggiorna_apparecchi(self):
        grid = self.ids.apparecchi_list
        grid.clear_widgets()
        if not self.apparecchi:
            grid.add_widget(Label(text="Nessun apparecchio.", size_hint_y=None, height=dp(30)))
            return
        for i, a in enumerate(self.apparecchi):
            testo = f"{a.tipologia} | {a.potenza}W | {a.installazione} | h={a.altezza_installazione}m | {a.accensione} | x{a.quantita}"
            row = BoxLayout(size_hint_y=None, height=dp(30), spacing=5)
            lbl = Label(text=testo, halign="left", size_hint_x=0.7, font_size=dp(12))
            btn_edit = Button(text="Mod", size_hint_x=0.15, font_size=dp(10))
            btn_edit.bind(on_press=lambda *args, idx=i: self.modifica_apparecchio(idx))
            btn_del = Button(text="X", size_hint_x=0.15, font_size=dp(10), background_color=(0.8, 0.2, 0.2, 1))
            btn_del.bind(on_press=lambda *args, idx=i: self.elimina_apparecchio(idx))
            row.add_widget(lbl)
            row.add_widget(btn_edit)
            row.add_widget(btn_del)
            grid.add_widget(row)

    def aggiorna_foto(self):
        grid = self.ids.foto_list
        grid.clear_widgets()
        if not self.foto_paths:
            grid.add_widget(Label(text="Nessuna foto.", size_hint_y=None, height=dp(30)))
            return
        for i, fp in enumerate(self.foto_paths):
            testo = os.path.basename(fp) if os.path.exists(fp) else f"[mancante] {os.path.basename(fp)}"
            row = BoxLayout(size_hint_y=None, height=dp(30), spacing=5)
            lbl = Label(text=testo, halign="left", size_hint_x=0.7, font_size=dp(11))
            btn_del = Button(text="X", size_hint_x=0.3, font_size=dp(10), background_color=(0.8, 0.2, 0.2, 1))
            btn_del.bind(on_press=lambda *args, idx=i: self.rimuovi_foto(idx))
            row.add_widget(lbl)
            row.add_widget(btn_del)
            grid.add_widget(row)

    def aggiungi_apparecchio(self):
        popup = FixturePopup(callback=self._on_fixture_added)
        popup.open()

    def _on_fixture_added(self, a):
        self.apparecchi.append(a)
        self.aggiorna_apparecchi()

    def modifica_apparecchio(self, idx):
        a = self.apparecchi[idx]
        popup = FixturePopup(apparecchio=a, callback=lambda nuovo: self._on_fixture_updated(idx, nuovo))
        popup.open()

    def _on_fixture_updated(self, idx, a):
        self.apparecchi[idx] = a
        self.aggiorna_apparecchi()

    def elimina_apparecchio(self, idx):
        self.apparecchi.pop(idx)
        self.aggiorna_apparecchi()

    def aggiungi_foto(self):
        from plyer import filechooser
        try:
            filechooser.open_file(on_selection=self._on_foto_scelta, filters=[("*.jpg", "*.jpeg", "*.png", "*.bmp")])
        except Exception:
            self._foto_da_filechooser_kivy()

    def _foto_da_filechooser_kivy(self):
        content = BoxLayout(orientation="vertical", padding=10, spacing=10)
        fc = FileChooserListView(filters=["*.jpg", "*.jpeg", "*.png", "*.bmp"])
        btn_row = BoxLayout(size_hint_y=0.15, spacing=10)
        btn_sel = Button(text="Seleziona")
        btn_cancel = Button(text="Annulla")
        popup = Popup(title="Seleziona foto", content=content, size_hint=(0.9, 0.8))
        btn_sel.bind(on_press=lambda *x: self._fc_selected(fc.selection, popup))
        btn_cancel.bind(on_press=popup.dismiss)
        btn_row.add_widget(btn_sel)
        btn_row.add_widget(btn_cancel)
        content.add_widget(fc)
        content.add_widget(btn_row)
        popup.open()

    def _fc_selected(self, selection, popup):
        popup.dismiss()
        if selection:
            self._on_foto_scelta([selection[0]])

    def _on_foto_scelta(self, paths):
        if paths and paths[0]:
            self.foto_paths.append(paths[0])
            self.aggiorna_foto()

    def scatta_foto(self):
        from plyer import camera
        try:
            foto_dir = App.get_running_app().foto_dir
            os.makedirs(foto_dir, exist_ok=True)
            path = os.path.join(foto_dir, f"foto_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
            camera.take_picture(path, self._on_foto_scattata)
        except Exception as e:
            App.get_running_app().mostra_messaggio("Errore", f"Fotocamera non disponibile:\n{str(e)}")

    def _on_foto_scattata(self, path):
        if path and os.path.exists(path):
            self.foto_paths.append(path)
            self.aggiorna_foto()

    def rimuovi_foto(self, idx):
        self.foto_paths.pop(idx)
        self.aggiorna_foto()

    def salva_stanza(self):
        id_locale = self.ids.id_locale.text.strip()
        if not id_locale:
            App.get_running_app().mostra_messaggio("Attenzione", "Inserisci l'ID del locale.")
            return
        stanza = Stanza(
            id_locale=id_locale,
            piano=self.ids.piano.text,
            altezza=self.ids.altezza.text,
            destinazione=self.ids.destinazione.text,
            controsoffitto=self.ids.controsoffitto.text,
            apparecchi=self.apparecchi,
            foto_paths=[p for p in self.foto_paths if os.path.exists(p)],
        )
        app = App.get_running_app()
        if self.index >= 0 and self.index < len(app.edificio.stanze):
            app.edificio.stanze[self.index] = stanza
        else:
            app.edificio.stanze.append(stanza)
        app.salva_json()
        self.manager.current = "stanze"


class ExportScreen(Screen):
    def on_enter(self, *args):
        if "info_label" in self.ids:
            self._populate_info()
        else:
            Clock.schedule_once(lambda *x: self.on_enter(), 0.2)

    def _populate_info(self):
        app = App.get_running_app()
        n_stanze = len(app.edificio.stanze)
        n_app = sum(sum(a.quantita for a in s.apparecchi) for s in app.edificio.stanze)
        n_foto = sum(len(s.foto_paths) for s in app.edificio.stanze)
        self.ids.info_label.text = (
            f"Edificio: {app.edificio.nome}\n"
            f"Stanze: {n_stanze}\n"
            f"Apparecchi: {n_app}\n"
            f"Foto: {n_foto}"
        )

    def esporta_excel(self):
        app = App.get_running_app()
        if not app.edificio.stanze:
            app.mostra_messaggio("Attenzione", "Nessuna stanza da esportare.")
            return

        def on_path(percorso):
            if not percorso:
                return
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

                wb = Workbook()
                ws1 = wb.active
                ws1.title = "Riepilogo"
                ws1.cell(row=1, column=1, value="RAPPORTO DI RILIEVO - RELAMPING").font = Font(bold=True, size=14)
                ws1.cell(row=3, column=1, value="Edificio:").font = Font(bold=True)
                ws1.cell(row=3, column=2, value=app.edificio.nome)
                ws1.cell(row=4, column=1, value="Indirizzo:").font = Font(bold=True)
                ws1.cell(row=4, column=2, value=app.edificio.indirizzo)
                ws1.cell(row=5, column=1, value="Data rilievo:").font = Font(bold=True)
                ws1.cell(row=5, column=2, value=app.edificio.data_rilievo)

                thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
                row_fill_odd = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

                headers1 = ["ID Locale", "Piano", "Altezza (m)", "Destinazione", "Controsoffitto", "N. App.", "Foto"]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                for ci, h in enumerate(headers1, 1):
                    c = ws1.cell(row=7, column=ci, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="center")

                for i, s in enumerate(app.edificio.stanze):
                    row = i + 8
                    n_app = sum(a.quantita for a in s.apparecchi)
                    vals = [s.id_locale, s.piano, s.altezza, s.destinazione, s.controsoffitto, n_app,
                            str(len(s.foto_paths)) if s.foto_paths else "No"]
                    for col, v in enumerate(vals, 1):
                        cell = ws1.cell(row=row, column=col, value=v)
                        cell.border = thin_border
                        if i % 2 == 1:
                            cell.fill = row_fill_odd

                ws2 = wb.create_sheet("Dettaglio Apparecchi")
                headers2 = ["ID Locale", "Tipologia", "Potenza (W)", "Installazione", "h_inst (m)", "Accensione", "Quantita"]
                for ci, h in enumerate(headers2, 1):
                    c = ws2.cell(row=1, column=ci, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.border = thin_border

                ri = 2
                for s in app.edificio.stanze:
                    for a in s.apparecchi:
                        vals = [s.id_locale, a.tipologia, a.potenza, a.installazione,
                                a.altezza_installazione, a.accensione, a.quantita]
                        for col, v in enumerate(vals, 1):
                            cell = ws2.cell(row=ri, column=col, value=v)
                            cell.border = thin_border
                        ri += 1

                ws1.column_dimensions["A"].width = 16
                ws1.column_dimensions["B"].width = 14
                ws1.column_dimensions["C"].width = 12
                ws1.column_dimensions["D"].width = 22
                ws1.column_dimensions["E"].width = 22
                ws1.column_dimensions["F"].width = 10
                ws1.column_dimensions["G"].width = 8
                for col in ["A", "B", "C", "D", "E", "F", "G"]:
                    ws2.column_dimensions[col].width = 16

                wb.save(percorso)

                if any(s.foto_paths for s in app.edificio.stanze):
                    zip_path = percorso.rsplit(".", 1)[0] + "_foto.zip"
                    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                        for s in app.edificio.stanze:
                            if s.foto_paths:
                                stanza_dir = "foto/" + s.id_locale.replace("/", "_").replace("\\", "_") + "/"
                                for fp in s.foto_paths:
                                    if os.path.exists(fp):
                                        zf.write(fp, stanza_dir + os.path.basename(fp))
                    app.mostra_messaggio("Esportato", f"Excel e foto salvati:\n{percorso}\n{zip_path}")
                else:
                    app.mostra_messaggio("Esportato", f"Excel salvato:\n{percorso}")
            except ImportError:
                app.mostra_messaggio("Errore", "openpyxl non installato.")
            except Exception as e:
                app.mostra_messaggio("Errore", f"Errore esportazione:\n{str(e)}")

        app._scegli_percorso("xlsx", on_path)

    def esporta_lsr(self):
        app = App.get_running_app()

        def on_path(percorso):
            if not percorso:
                return
            try:
                data = app.edificio.to_dict()
                with tempfile.TemporaryDirectory() as tmpdir:
                    with open(os.path.join(tmpdir, "rilievo.json"), "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=2, ensure_ascii=False)
                    foto_dir = os.path.join(tmpdir, "foto")
                    os.makedirs(foto_dir, exist_ok=True)
                    for s in app.edificio.stanze:
                        for idx, fp in enumerate(s.foto_paths):
                            if os.path.exists(fp):
                                base = os.path.basename(fp)
                                name, ext = os.path.splitext(base)
                                shutil.copy2(fp, os.path.join(foto_dir, f"{name}_{idx}{ext}"))
                    with zipfile.ZipFile(percorso, "w", zipfile.ZIP_DEFLATED) as zf:
                        for root, _, files in os.walk(tmpdir):
                            for fn in files:
                                zf.write(os.path.join(root, fn), os.path.relpath(os.path.join(root, fn), tmpdir))
                app.mostra_messaggio("Salvato", f"Progetto salvato:\n{percorso}")
            except Exception as e:
                app.mostra_messaggio("Errore", f"Salvataggio fallito:\n{str(e)}")

        app._scegli_percorso("lsr", on_path)


class LusorApp(App):
    def build(self):
        self.icon = "icon.png"
        self.edificio = Edificio()
        self.data_dir = self._get_data_dir()
        os.makedirs(self.data_dir, exist_ok=True)
        self.foto_dir = os.path.join(self.data_dir, "foto")
        os.makedirs(self.foto_dir, exist_ok=True)
        self.json_path = os.path.join(self.data_dir, "rilievo.json")
        self.lsr_temp_dir = None
        sm = ScreenManager()
        sm.add_widget(HomeScreen(name="home"))
        sm.add_widget(RoomListScreen(name="stanze"))
        sm.add_widget(RoomDetailScreen(name="stanza_detail"))
        sm.add_widget(ExportScreen(name="export"))
        # Popola campi dopo il primo frame
        Clock.schedule_once(lambda *x: sm.get_screen("home").on_enter(), 0.3)
        return sm

    def _get_data_dir(self):
        if platform == "android":
            from android.storage import app_storage_path
            return app_storage_path()
        return os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "Lusor")

    def carica_json(self):
        try:
            if os.path.exists(self.json_path):
                with open(self.json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.edificio = Edificio.from_dict(data)
        except Exception as e:
            print(f"Errore caricamento: {e}")

    def salva_json(self):
        try:
            with open(self.json_path, "w", encoding="utf-8") as f:
                json.dump(self.edificio.to_dict(), f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Errore salvataggio: {e}")

    def _scegli_percorso(self, ext, callback):
        from plyer import filechooser
        ext_map = {"xlsx": "*.xlsx", "lsr": "*.lsr"}
        try:
            filechooser.save_file(on_selection=lambda s: callback(s[0] if s else None),
                                  filters=[ext_map.get(ext, "*.*")])
        except Exception:
            self._percorso_fallback(ext, callback)

    def _percorso_fallback(self, ext, callback):
        default_name = f"rilievo_{self.edificio.nome.replace(' ', '_')}.{ext}"
        from kivy.uix.textinput import TextInput
        content = BoxLayout(orientation="vertical", padding=15, spacing=15)
        ti = TextInput(text=os.path.join(self.data_dir, default_name), multiline=False)
        btn_row = BoxLayout(spacing=10, size_hint_y=0.4)
        btn_save = Button(text="Salva")
        btn_cancel = Button(text="Annulla")
        popup = Popup(title="Salva", content=content, size_hint=(0.8, 0.35))

        def on_save(*args):
            percorso = ti.text.strip()
            popup.dismiss()
            if percorso:
                callback(percorso)
            else:
                callback(None)

        btn_save.bind(on_press=on_save)
        btn_cancel.bind(on_press=lambda *x: [popup.dismiss(), callback(None)])
        btn_row.add_widget(btn_save)
        btn_row.add_widget(btn_cancel)
        content.add_widget(ti)
        content.add_widget(btn_row)
        popup.open()

    def mostra_messaggio(self, titolo, testo):
        popup = MessaggioPopup(titolo, testo)
        popup.open()


if __name__ == "__main__":
    LusorApp().run()
